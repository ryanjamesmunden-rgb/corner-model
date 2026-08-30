"""Rebuild Corner_EV_Ledger.xlsx — the weekly EV ledger — from the rows in ROWS below.

    python3 tools/build_ev_ledger.py        # writes Corner_EV_Ledger.xlsx at the repo root

The workbook is a rebuild of a handwritten ledger, so its shape is NOT free: the columns,
the yellow-cells-only convention and the Summary metrics all match the sheet that already
existed. Keep them.

VERIFYING A CHANGE. LibreOffice cannot load a file in this sandbox, so `recalc.py` is not
available and the workbook ships with formulas but no cached values (Excel and Sheets
recalculate on open, so this is invisible to a reader). That removes the usual safety net,
and the replacement is the source ledger's own summary — 55 fixtures, 21 leagues, 51 priced,
average EV 9.9%, 28 positive, 22 negative, best 119.3% on Kortrijk v Charleroi, 1 won,
100% strike. Re-derive those in Python from ROWS and compare after any edit. That check is
what caught the staking block landing on top of Won/Lost/Void, and `Best EV fixture`
matching against the positive-bet count instead of the best EV — two edits that produced a
workbook which opened perfectly and reported the wrong thing.

NEVER hardcode a Summary row number. Sections are laid out once into ADDR and formulas are
written with {Label} tokens substituted from it; an assert catches any drift.
"""
import os
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Corner_EV_Ledger.xlsx")
A = "Arial"
BLUE = Font(name=A, size=10, color="0000FF")
BLACK = Font(name=A, size=10)
BOLD = Font(name=A, size=10, bold=True)
HEAD = Font(name=A, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=A, size=14, bold=True)
SUB = Font(name=A, size=10, italic=True, color="595959")
NOTE = Font(name=A, size=9, color="595959")
YEL = PatternFill("solid", fgColor="FFFF00")
HF = PatternFill("solid", fgColor="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

D = dt.date
# Date, League, Fixture, Backing, Proj, Line, Fair, Your odds, Closing, Corners, Result, Notes
ROWS = [
 (D(2026,8,25),"Brazil Serie B","Atletico GO v Botafogo SP","Atletico GO",7.5,"6+",1.47,1.66,None,None,"W","Price drifted in: taken 1.72, now 1.66"),
 (D(2026,8,28),"Spain La Liga","Alaves v Villarreal","Alaves",6.1,"5+",1.49,1.50,None,None,None,"6+ also quoted 1.90"),
 (D(2026,8,28),"Portugal Primeira Liga","Rio Ave v Sporting CP","Sporting CP",6.4,"5+",1.43,1.22,None,None,None,None),
 (D(2026,8,28),"Belgium Pro League","Genk v Beveren","Genk",None,"6+",2.16,2.00,None,None,None,"No projection or line written yet"),
 (D(2026,8,29),"Italy Serie A","Juventus v Parma","Juventus",7.1,"6+",1.53,1.36,None,None,None,None),
 (D(2026,8,29),"Italy Serie A","Fiorentina v Fros","Fiorentina",8.5,"8+",1.78,3.75,None,None,None,"Also 3+ FHC 1.90, 3+ SHC 1.72. Opponent conceded 9 first-half corners v Juventus last week"),
 (D(2026,8,29),"France Ligue 1","Lyon v Le Havre","Lyon",6.5,"5+",1.41,1.28,None,None,None,"7+ also quoted 1.83"),
 (D(2026,8,29),"Germany Bundesliga","Borussia Dortmund v Hamburger SV","Borussia Dortmund",7.2,"6+",1.51,1.44,None,None,None,None),
 (D(2026,8,29),"Germany Bundesliga","RB Leipzig v Monchengladbach","RB Leipzig",5.7,"5+",1.60,1.36,None,None,None,None),
 (D(2026,8,29),"Germany Bundesliga 2","Energie Cottbus v SpVgg","Energie Cottbus",6.2,"5+",1.58,1.72,None,None,None,"Projection revised 5.8 -> 6.2. Marked OVERS!"),
 (D(2026,8,29),"Germany Bundesliga 2","Energie Cottbus v SpVgg","Over 9.5 Asian Corners",None,None,None,1.83,None,None,None,"Second bet on the same fixture — no fair odds written, so no EV"),
 (D(2026,8,29),"England Championship","Cardiff v Sheff Utd","Cardiff",7.2,"6+",1.51,1.57,None,None,None,"Cardiff 8.0 won / Sheff Utd 6.9 conceded"),
 (D(2026,8,29),"England League One","Sheff Wed v Bromley","Sheff Wed",7.9,"7+",1.63,1.90,None,None,None,"Sheff Wed 7.5 won / Bromley 8.5 conceded"),
 (D(2026,8,29),"England League One","Stockport v Wycombe","Stockport",6.9,"6+",1.60,1.57,None,None,None,"Stockport 6.3 won / Wycombe 5.7 conceded"),
 (D(2026,8,29),"England League One","Leyton Orient v Barnsley","Leyton Orient",6.1,"5+",1.49,1.57,None,None,None,"Orient 4.6 won / Barnsley 6.1 conceded"),
 (D(2026,8,29),"England League Two","York v Exeter","Exeter",5.5,"4+",1.36,1.53,None,None,None,"Date assumed 29/08. 5+ quoted 2.10, ticked. York 5.5 / Exeter 6.0"),
 (D(2026,8,29),"England League Two","Rotherham v Chesterfield","Chesterfield",6.4,"5+",1.43,1.72,None,None,None,"Date assumed 29/08. Rotherham 7.0 / Chesterfield 5.7"),
 (D(2026,8,29),"England League Two","Colchester v Rochdale","Colchester",5.6,"5+",1.63,1.66,None,None,None,"Date assumed 29/08. Colchester 5.2 / Rochdale 6.0"),
 (D(2026,8,29),"England League Two","Newport County v Tranmere","Newport County",6.2,"5+",1.47,1.72,None,None,None,"Date assumed 29/08. Market N/A. Newport 5.15 / Tranmere 5.9"),
 (D(2026,8,29),"Belgium Pro League","Kortrijk v Charleroi","Charleroi",7.6,"7+",1.71,3.75,None,None,None,None),
 (D(2026,8,29),"Belgium Pro League","Lommel v Cercle Brugge","Cercle Brugge",7.7,"7+",1.70,1.61,None,None,None,"Home arrow noted against Lommel"),
 (D(2026,8,29),"Belgium Pro League","OH Leuven v Standard Liege","Standard Liege",6.4,"5+",1.44,1.66,None,None,None,None),
 (D(2026,8,29),"Scotland Premiership","Celtic v Falkirk","Celtic",6.9,"6+",1.60,1.22,None,None,None,"Crossed out - no value at 1.22"),
 (D(2026,8,29),"Turkey Super Lig","Galatasaray v Goztepe","Galatasaray",6.7,"6+",1.65,1.66,None,None,None,None),
 (D(2026,8,29),"Japan J League","Nagoya Grampus v Fagiano Okayama",None,5.8,"5+",1.57,1.66,None,None,None,None),
 (D(2026,8,29),"Japan J League","Vissel Kobe v Cerezo Osaka",None,6.6,"6+",1.69,2.10,None,None,None,"Vissel Kobe havent won in 4 h2hs"),
 (D(2026,8,29),"Japan J League","Urawa v Yokohama",None,5.8,"5+",1.58,1.53,None,None,None,None),
 (D(2026,8,30),"Brazil Serie A","Sao Paulo v RB Bragantino","Sao Paulo",7.7,"7+",1.68,2.30,None,None,None,None),
 (D(2026,8,30),"Brazil Serie A","Mirassol v Palmeiras","Mirassol",6.6,"6+",1.68,2.25,None,None,None,None),
 (D(2026,8,30),"Brazil Serie A","Flamengo v Botafogo","Flamengo",6.1,"5+",1.49,1.22,None,None,None,None),
 (D(2026,8,30),"Brazil Serie A","Gremio v Chapecoense","Gremio",6.5,"5+",1.41,1.33,None,None,None,None),
 (D(2026,8,30),"England Premier League","Leeds v Brentford","Leeds",6.5,"6+",1.70,2.37,None,None,None,"Leeds 4.3 won / Brentford 7.3 conceded"),
 (D(2026,8,30),"England Premier League","Liverpool v Nottingham Forest","Liverpool",7.0,"6+",1.56,1.50,None,None,None,"Swapped round from Forest v Liverpool. Liverpool 6.9 won / Forest 6.4 conceded"),
 (D(2026,8,30),"Portugal Primeira Liga","Nacional v Estrela","Nacional",5.9,"5+",1.55,None,None,None,None,"Marked AWAY"),
 (D(2026,8,30),"Belgium Pro League","Gent v Club Brugge","Club Brugge",6.1,"5+",1.49,1.40,None,None,None,None),
 (D(2026,8,30),"Germany Bundesliga 2","St Pauli v 1. FC Kaiserslautern","St Pauli",6.9,"6+",1.58,1.61,None,None,None,None),
 (D(2026,8,30),"USA MLS","Minnesota v Orlando","Minnesota",6.5,"5+",1.42,1.36,None,None,None,None),
 (D(2026,8,30),"USA MLS","Nashville v Cincinnati",None,6.0,"5+",1.51,1.44,None,None,None,None),
 (D(2026,8,30),"USA MLS","Colorado Rapids v Real Salt Lake",None,6.9,"6+",1.60,1.90,None,None,None,"1-0 or 2-0 in last 4 home games. under goals?"),
 (D(2026,8,30),"Greece Super League","Aris v OFI",None,6.0,"5+",1.51,None,None,None,None,None),
 (D(2026,8,30),"Argentina Liga Profesional","Independiente v Gimnasia","Independiente",7.5,"6+",1.46,1.83,None,None,None,"Won 6+ in 10 of 11 at home"),
 (D(2026,8,30),"Argentina Liga Profesional","Union Santa Fe v Sarmiento",None,7.7,"7+",1.68,2.00,None,None,None,None),
 (D(2026,8,30),"Argentina Liga Profesional","Argentinos Jrs v Aldosivi",None,6.7,"6+",1.66,1.57,None,None,None,"Crossed out at 1.57. 7+ quoted 2.00"),
 (D(2026,8,30),"Argentina Liga Profesional","Atletico Tucuman v Belgrano",None,5.8,"5+",1.58,1.83,None,None,None,"Ticked"),
 (D(2026,8,30),"Argentina Liga Profesional","Rosario Central v Gimnasia LP",None,6.2,"5+",1.47,1.33,None,None,None,"7+ quoted 2.10, crossed out. 4+ quoted 1.40"),
 (D(2026,8,30),"Argentina Liga Profesional","Independiente Rivadavia v Racing Club",None,5.2,"4+",1.42,None,None,None,None,"5+ quoted 1.83"),
 (D(2026,8,30),"Argentina Liga Profesional","Estudiantes LP v Newells Old Boys","Estudiantes LP",6.6,"6+",1.70,2.87,None,None,None,"5+ quoted 2.00, marked green"),
 (D(2026,8,30),"Argentina Liga Profesional","Talleres v Central Cordoba (Santiago)",None,5.6,"5+",1.64,1.57,None,None,None,"6+ quoted 2.10, crossed out"),
 (D(2026,8,31),"Turkey Super Lig","Amed v Trabzonspor","Amed",5.3,"4+",1.40,None,None,None,None,None),
 (D(2026,8,31),"USA MLS","St Louis City v FC Dallas","St Louis City",6.5,"6+",1.71,1.66,None,None,None,None),
 (None,"Brazil Serie B","Novorizontino v Recife",None,7.4,"6+",1.48,1.40,None,None,None,"No date written (round 24/25)"),
 (None,"Brazil Serie B","Nautico Recife v Athletic",None,7.0,"6+",1.56,1.80,None,None,None,"No date written (round 24/25)"),
 (None,"Brazil Serie B","Villa Nova v Ceara",None,6.7,"6+",1.64,1.66,None,None,None,"No date written (round 24/25)"),
 (None,"Switzerland Super League","Young Boys v Basel",None,6.5,"5+",1.42,1.33,None,None,None,"No date written (round 5)"),
 (None,"Austria Bundesliga","LASK Linz v Altach",None,6.8,"6+",1.64,1.90,None,None,None,"No date written (round 5)"),
]

wb = Workbook()

# ------------------------------------------------------------------ Fixtures
fx = wb.active
fx.title = "Fixtures"
COLS = [("Date",11),("League",22),("Fixture",34),("Backing",22),("Proj",7),("Line",7),
        ("Fair odds",10),("Your odds",10),("EV %",9),("Closing",9),("CLV %",9),
        ("Corners",9),("Result",8),("Notes",58),("Stake £",9),("_first",6)]
for i,(n,w) in enumerate(COLS, start=1):
    c = fx.cell(row=1, column=i, value=n)
    c.font, c.fill, c.border = HEAD, HF, BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    fx.column_dimensions[get_column_letter(i)].width = w
fx.row_dimensions[1].height = 26

FIRST, SPARE = 2, 45
LAST = FIRST + len(ROWS) - 1
END = LAST + SPARE

for r in range(FIRST, END + 1):
    d = ROWS[r - FIRST] if r <= LAST else None
    def put(col, val, font=BLACK, fmt=None, fill=None):
        c = fx.cell(row=r, column=col, value=val)
        c.font, c.border = font, BOX
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c
    put(1, d[0] if d else None, BLUE, "yyyy-mm-dd")
    put(2, d[1] if d else None, BLUE)
    put(3, d[2] if d else None, BLUE)
    put(4, d[3] if d else None, BLUE)
    put(5, d[4] if d else None, BLUE, "0.0")
    put(6, d[5] if d else None, BLUE)
    put(7, d[6] if d else None, BLUE, "0.00")
    put(8, d[7] if d else None, BLUE, "0.00", YEL)            # Your odds
    put(9, f'=IF(OR(G{r}="",H{r}=""),"",H{r}/G{r}-1)', BLACK, "0.0%")
    put(10, d[8] if d else None, BLUE, "0.00", YEL)           # Closing
    put(11, f'=IF(OR(H{r}="",J{r}=""),"",H{r}/J{r}-1)', BLACK, "0.0%")
    put(12, d[9] if d else None, BLUE, "0", YEL)              # Corners
    put(13, d[10] if d else None, BLUE, None, YEL)            # Result
    put(14, d[11] if d else None, BLUE).alignment = Alignment(wrap_text=True, vertical="top")
    put(15, f'=IF(OR(I{r}="",I{r}<=0,H{r}<=1),"",'
            f'MIN(I{r}/(H{r}-1),{{Max stake (% of bank)}})*{{Kelly fraction}}*{{Bankroll (£)}})',
        BLACK, "£#,##0.00")
    put(16, f'=IF(B{r}="",0,IF(COUNTIF($B$2:B{r},B{r})=1,1,0))', BLACK, "0")

fx.freeze_panes = "A2"
fx.auto_filter.ref = f"A1:O{END}"
fx.column_dimensions["P"].hidden = True

dv = DataValidation(type="list", formula1='"W,L,V"', allow_blank=True, showDropDown=False)
dv.error = "Pick W (won), L (lost) or V (void)."
fx.add_data_validation(dv)
dv.add(f"M{FIRST}:M{END}")

rng = f"I{FIRST}:I{END}"
fx.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
    font=Font(name=A, size=10, bold=True, color="006100"), fill=PatternFill("solid", fgColor="C6EFCE")))
fx.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
    font=Font(name=A, size=10, color="9C0006"), fill=PatternFill("solid", fgColor="FFC7CE")))
body = f"A{FIRST}:O{END}"
fx.conditional_formatting.add(body, FormulaRule(formula=[f'$M{FIRST}="W"'],
    fill=PatternFill("solid", fgColor="E2EFDA")))
fx.conditional_formatting.add(body, FormulaRule(formula=[f'$M{FIRST}="L"'],
    fill=PatternFill("solid", fgColor="FCE4EC")))

for col, txt in [
    ("H", "The only price that matters: what you can ACTUALLY get. EV compares it to Fair odds."),
    ("I", "= Your odds / Fair odds - 1. Positive means the price beats your model's own number.\n\n"
          "This is EV per £1 staked, not just 'edge', because your Fair odds are the break-even "
          "price: EV is exactly zero when the two match."),
    ("J", "Closing price, filled in after the market settles. CLV is the only feedback that "
          "arrives fast enough to be useful — results take hundreds of bets to say anything, "
          "closing line value tells you within a week whether you are beating the market."),
    ("O", "Quarter-Kelly stake from the EV and your bankroll, capped. Set bankroll, fraction and "
          "cap on the Summary tab. Blank when there is no edge — nothing to stake."),
]:
    fx[f"{col}1"].comment = Comment(txt, "Corner Model", height=170, width=330)

# ------------------------------------------------------------------ Summary
sm = wb.create_sheet("Summary")
sm["A1"] = "Corner ledger"
sm["A1"].font = TITLE
sm["A2"] = "Every figure recalculates from the Fixtures tab. Type only in the yellow cells."
sm["A2"].font = SUB
L, R = f"$I${FIRST}:$I${END}", f"$M${FIRST}:$M${END}"
F = "Fixtures!"
SECTIONS = [
    ("Coverage", [
        ("Fixtures", f'=COUNTA({F}$C${FIRST}:$C${END})', "0", ""),
        ("Leagues", f'=SUM({F}$P${FIRST}:$P${END})', "0", ""),
        ("Priced up", f'=COUNT({F}$H${FIRST}:$H${END})', "0", "Rows where you've entered your odds"),
        ("Still to price", f'={{Fixtures}}-{{Priced up}}', "0", "Has a fixture but no price yet"),
    ]),
    ("Value", [
        ("Average EV", f'=IFERROR(AVERAGE({F}{L}),"")', "0.0%", "Across priced rows only"),
        ("Positive EV bets", f'=COUNTIF({F}{L},">0")', "0", ""),
        ("Negative EV bets", f'=COUNTIF({F}{L},"<0")', "0", "Your reds"),
        ("Best EV on the card", f'=IFERROR(MAX({F}{L}),"")', "0.0%", ""),
        ("Best EV fixture",
         f'=IFERROR(INDEX({F}$C${FIRST}:$C${END},MATCH({{Best EV on the card}},{F}{L},0)),"")', None, ""),
    ]),
    ("Closing line value", [
        ("Average CLV", f'=IFERROR(AVERAGE({F}$K${FIRST}:$K${END}),"needs closing prices")', "0.0%", ""),
        ("Beat the close", f'=COUNTIF({F}$K${FIRST}:$K${END},">0")', "0", ""),
        ("Beaten by the close", f'=COUNTIF({F}$K${FIRST}:$K${END},"<0")', "0", ""),
    ]),
    ("Results", [
        ("Won", f'=COUNTIF({F}{R},"W")', "0", ""),
        ("Lost", f'=COUNTIF({F}{R},"L")', "0", ""),
        ("Void", f'=COUNTIF({F}{R},"V")', "0", "Stake returned — excluded from strike rate"),
        ("Settled", f'={{Won}}+{{Lost}}', "0", ""),
        ("Strike rate", f'=IFERROR({{Won}}/{{Settled}},"")', "0.0%", "Voids excluded"),
    ]),
    ("Staking", [
        ("Bankroll (£)", 1000, "£#,##0", "Drives the Stake column on Fixtures."),
        ("Kelly fraction", 0.25, "0.00",
         "Quarter Kelly. Full Kelly assumes your fair odds are exactly right; they are your own "
         "model's, and its edge is not yet measured, so bet a fraction of what Kelly says."),
        ("Max stake (% of bank)", 0.02, "0.0%",
         "Hard cap, so one short-priced line cannot take the bank."),
    ]),
]

# Lay the sections out on paper first, so every cross-reference is written from the row a
# label ACTUALLY lands on. Hardcoding them put the staking block on top of Won/Lost/Void and
# pointed "Best EV fixture" at the positive-bet count — both of which produce a workbook that
# opens clean and reports the wrong thing.
ADDR, _r = {}, 4
for _h, _items in SECTIONS:
    _r += 1
    for _lab, *_ in _items:
        ADDR[_lab] = _r
        _r += 1
    _r += 1

def sub(v):
    """Replace {Label} with the B-column address that label was laid out at."""
    if isinstance(v, str) and "{" in v:
        for k, row in ADDR.items():
            v = v.replace("{" + k + "}", f"$B${row}")
    return v

for _r2 in range(FIRST, END + 1):
    _c = fx.cell(row=_r2, column=15)
    _c.value = sub(_c.value).replace("$B$", "Summary!$B$")
r = 4
for heading, items in SECTIONS:
    sm[f"A{r}"] = heading
    sm[f"A{r}"].font = Font(name=A, size=11, bold=True, color="1F3864")
    r += 1
    for label, formula, fmt, note in items:
        assert ADDR[label] == r, f"{label}: laid out at {ADDR[label]}, writing at {r}"
        sm[f"A{r}"], sm[f"A{r}"].font = label, BOLD
        c = sm[f"B{r}"]
        c.value, c.border = sub(formula), BOX
        # a typed-in setting is a blue input on yellow; everything else is a formula
        if isinstance(formula, str):
            c.font = BLACK
        else:
            c.font, c.fill = BLUE, YEL
        if fmt: c.number_format = fmt
        sm[f"C{r}"], sm[f"C{r}"].font = note, NOTE
        sm[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

HOW = [
    ("How to use", [
        ("Yellow cells", "Your odds, Closing, Corners, Result on Fixtures; bankroll and staking here."),
        ("EV %", "Your odds / Fair odds - 1. Green above zero, red below."),
        ("CLV %", "Your odds / Closing - 1. Whether you beat the close."),
        ("Result", "Pick W, L or V from the dropdown. The row tints green or red."),
        ("Stake £", "Quarter-Kelly from the EV, capped. Blank when there is no edge."),
        ("Filters", "Row 1 has filter arrows — narrow by date or league."),
        ("Example", "Rotherham v Chesterfield: Fair 1.43, type 1.72 into Your odds -> EV shows +20.3%."),
    ]),
    ("Carried over from the handwritten pages", [
        ("League Two dates", "Not written on the page — assumed Sat 29 Aug."),
        ("Five undated fixtures", "Three Serie B, Young Boys v Basel, LASK v Altach — Date left blank."),
        ("Different-line prices", "Where your market price was for a different line than the fair odds, "
                                  "it sits in Notes, not in Your odds."),
        ("Fair odds", "Your own figures, copied as written — not recalculated."),
    ]),
    ("One thing to watch on UNDER lines", [
        ("Voids", "Your model's published price is already void-adjusted: odds = 1 + lose/win, NOT 1/prob. "
                  "On an under, landing exactly on the line returns the stake."),
        ("Why it matters", "Bahia under 7 shows 1.30, but 1/0.692 = 1.45. If you ever compute a fair price "
                           "from a percentage yourself, use the void-adjusted form or every under is wrong."),
        ("Here", "Not an issue — Fair odds are copied from the model, which has already done this."),
    ]),
]
r += 1
for heading, items in HOW:
    sm[f"A{r}"] = heading
    sm[f"A{r}"].font = Font(name=A, size=11, bold=True, color="1F3864")
    r += 1
    for k, v in items:
        sm[f"A{r}"], sm[f"A{r}"].font = k, BOLD
        sm[f"B{r}"], sm[f"B{r}"].font = v, Font(name=A, size=10)
        sm.merge_cells(f"B{r}:E{r}")
        sm[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        sm.row_dimensions[r].height = 28
        r += 1
    r += 1

sm.column_dimensions["A"].width = 26
sm.column_dimensions["B"].width = 24
sm.column_dimensions["C"].width = 46
for col in "DE":
    sm.column_dimensions[col].width = 20

wb.save(OUT)
print("wrote", OUT, "| fixture rows:", len(ROWS))
